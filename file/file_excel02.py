from openpyxl import load_workbook

# 读取xlsx文件
wb = load_workbook('TestFile/test_xlsx.xlsx')
sheet = wb.worksheets[0]

# 获取第n行第n列的单元格,行列都是从1开始的
"""
cell = sheet.cell(1, 1)

print(cell.value)
print(cell.style)
print(cell.font)
print(cell.alignment)
"""

# 获取某个单元格
"""
c1 = sheet['A1']
print(c1.value)
c2 = sheet['A4']
print(c2.value)
"""

# 获取第N行所有的单元格
"""
for cell in sheet[1]:
    print(cell.value)

"""

# 获取所有行的数据
"""
for row in sheet.rows:
    print(row[0].value)  # 第一列的数据
"""

# 获取所有列

for columns in sheet.columns:
    print(columns[1].value)  # 第二行的数据

# 读取合并单元格
sheet1 = wb.worksheets[1]
c1 = sheet1.cell(5, 4)
print(c1)  # <Cell '合约框架基础信息'.D5>
print(c1.value)  # 佣金设置
c2 = sheet1.cell(5, 5)
print(c2)  # <MergedCell '合约框架基础信息'.E5>
print(c2.value)  # None


