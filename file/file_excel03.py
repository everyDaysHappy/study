from openpyxl import load_workbook, workbook, styles
from openpyxl.styles import Border, Alignment, Side, Font, PatternFill, GradientFill

# 在源文件的基础上修改


"""
wb = load_workbook('TestFile/test_xlsx.xlsx')
sheet = wb['TestSheet']
cell = sheet.cell(1, 1)
cell.value = 'excel练习'
# excel文件保存
wb.save('TestFile/test_xlsx.xlsx')
"""

# 新建文件写内容
"""
# 创建excel文件，自动创建Sheet
wb = workbook.Workbook()

sheet = wb.worksheets[0]

cell = sheet.cell(1, 1)
cell.value = '新文件'

wb.save('TestFile/p1.xlsx')
"""

# 修改sheet
# wb = workbook.Workbook()

# 修改sheet名字
"""
sheet = wb.worksheets[0]
sheet.title = '修改sheet名称'
wb.save('')
"""

# 创建sheet并设置sheet颜色
"""
sheet = wb.create_sheet('创建sheet', 0)
sheet.sheet_properties.tabColor = '1072BA'
wb.save('test.xlsx')
"""

# 设置默认打开的sheet
"""
wb.active = 0
wb.save('test.xlsx')
"""

# 拷贝sheet
"""
sheet = wb.create_sheet('拷贝sheet1')
sheet.sheet_properties.tabColor = '1072BA'

new_sheet = wb.copy_worksheet(wb['Sheet'])
new_sheet.title = '拷贝sheet2'

wb.save('test.xlsx')
"""

# 修改单元格
"""
wb = load_workbook('test.xlsx')
sheet = wb.worksheets[0]
"""
# 修改单元格的值
"""
sheet['B3'] = '开始'
wb.save()
"""

# 批量修改
"""
cell_list = sheet.['A3':'E3']
for row in cell_list:
    for cell in row:
        cell.value = '666'
wb.save('test.xlsx')
"""

# 对其方式
"""
cell = sheet.cell(1, 1)
print(cell.value)
# horizontal 水平方向对齐方式
# vertical 垂直方向对齐方式
# text_rotation 旋转角度
# wrap_text 是否自动换行
cell.alignment = Alignment(horizontal='center', vertical='distributed', text_rotation=45, wrap_text=True)
wb.save('test.xlsx')
"""

# 边框
"""
cell = sheet.cell(2, 2)
cell.border = Border(
    top=Side(style='thin', color='FFB6C1'),
    bottom=Side(style='dashed', color='9932CC'),
    left=Side(style='dashed', color='9932CC'),
    right=Side(style='dashed', color='9932CC'),
    diagonal=Side(style='thin', color='483D8B'),  # 对角线
    diagonalUp=True,  # 左下-右上
    diagonalDown=True  # 左上-右下
)
wb.save('test.xlsx')
"""

# 字体
"""
cell = sheet.cell(3, 3)
cell.value = '字体'
cell.font = Font(name='微软雅黑', size=45, color='ff0000', underline='single')
"""
# 背景色
# cell.fill = PatternFill('soild', fgColor='99ccff')

# 渐变背景色
# cell.fill = GradientFill('linear', stop=('FFFFFF', '99ccff', '000000'))

# 宽高(索引从1开始)
# sheet.row_dimensions['1'].height = 50
# sheet.column_dimensions['2'].width = 100

# 合并单元格
# sheet.merge_cells('B3:D6')
# sheet.unmerge_cells() 拆分单元格

# 写入公式
# sheet['D1'] = '公式'
# sheet['D2'] = '=B2*C2'
# wb.save('test.xlsx')

# 删除
# idx 要删除的索引位置
# amount 从索引位置开始要删除的个数
# sheet.delete_rows(idx=1, amount=20)
# sheet.delete_cols(idx=1, amount=3)

# 插入
"""
sheet.insert_cols(idx=1, amount=2)
sheet.insert_rows(idx=1, amount=2)
"""

# 循环写内容
"""
sheet1 = wb['Sheet']
cell_range = sheet1['A1:C2']
for row in cell_range:
    for Cell in row:
        Cell.value = 'xx'

for row in sheet1.iter_rows(min_row=5, min_col=1, max_col=7, max_row=10):
    for cell in row:
        cell.value = 'oo'
wb.save('test.xlsx')
"""

wb = load_workbook('TestFile/p1.xlsx')
sheet = wb.worksheets[0]

# 移动
sheet.move_range('H2:J10', rows=-1, cols=15)

# 打印区域
sheet.print_area = 'A1:D200'

# 打印时，每个页面的固定表头
sheet.print_title_cols = 'A:D'
sheet.print_title_rows = '1:3'

wb.save('TestFile/p1.xlsx')
