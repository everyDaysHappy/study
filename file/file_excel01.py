from openpyxl import load_workbook

# 读取xlsx文件
wb = load_workbook('TestFile/test_xlsx.xlsx')

# sheet相关操作

# 1.获取sheet名称
print(wb.sheetnames)

# 2.选择sheet页
sheet = wb['day1收盘价']

cell = sheet.cell(2, 2)
print(cell.value)

# 3.选择sheet，基于索引位置
sheet1 = wb.worksheets[1]
cell1 = sheet1.cell(1, 3)
print(cell1.value)

# 4.循环所有的sheet
for name in wb.sheetnames:
    Sheet = wb[name]
    Cell = Sheet.cell(1, 1)
    print(Cell.value)

# for sheet in wb.worksheets:
#     cell = sheet.cell(1, 1)
#     print(cell.value)

# for sheet in wb:
#     cell = sheet.cell(1, 1)
#     print(cell.value)
